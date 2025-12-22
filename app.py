from flask import Flask, request, send_file, render_template, redirect, url_for
from weasyprint import HTML, CSS
from pypdf import PdfWriter
import io
import csv
import zipfile
import tempfile
import shutil
import os
from datetime import datetime
import threading
import uuid
import openpyxl
from openpyxl import Workbook, load_workbook
import pathlib


# ... (dopo gli import, prima di app = Flask...)

# Configurazione percorsi di archiviazione (MODIFICA QUESTI PERCORSI)
PATH_ARCHIVIO_1 = "C://xampp/htdocs//generatorediplomi6.8//Copia_1"
PATH_ARCHIVIO_2 = "C://xampp/htdocs//generatorediplomi6.8//Copia_2"
PATH_EXCEL_REGISTRO = "C://xampp/htdocs//generatorediplomi6.8//Registri"

# Assicurati che le cartelle esistano all'avvio
for p in [PATH_ARCHIVIO_1, PATH_ARCHIVIO_2, PATH_EXCEL_REGISTRO]:
    os.makedirs(p, exist_ok=True)


def format_place_name(place_str):
    """
    Formatta una stringa di luogo (Comune, Provincia, Stato):
    - Title-case tutte le parole.
    - Converte in minuscolo le preposizioni, gli articoli e le forme elise ('d', 'l').
    - Gestisce correttamente la capitalizzazione dei nomi propri con apostrofo (es. L'Aquila).
    """
    if not place_str:
        return ""

    # Set di preposizioni/articoli comuni da mantenere in minuscolo
    PREPOSITIONS = {
        'di', 'del', 'della', 'degli', 'dei', 'de', 'da', 'dal', 
        'dalla', 'dai', 'dagli', 'su', 'sul', 'sulla', 'sui', 
        'sugli', 'a', 'al', 'alla', 'ai', 'agli', 'in', 'nel', 
        'nella', 'nei', 'negli', 'per', 'con', 'e', 'il', 'lo', 'la', 'gli', 'le', 
        "d'", "l'","de'",'val','meno',"all'"
    }

    # Converto l'intera stringa in minuscolo per processare le parole
    words = place_str.lower().split()
    final_formatted_words = []
    
    for i, word in enumerate(words):
        
        formatted_word = word
        
        # 1. Gestione di TUTTE le parole che sono solo preposizioni/articoli e non sono la prima
        if word in PREPOSITIONS and i > 0:
            final_formatted_words.append(word.lower()) 
            continue
            
        # 2. Gestione di Elisioni complesse con apostrofo ('d'America', 'L'Aquila')
        if word.find('\'') != -1:
            index = word.find('\'')
            
            # Parte prima dell'apostrofo (d, l, ecc.)
            prefix = word[:index+1]
            
            # Parte dopo l'apostrofo (america, aquila, ecc.)
            suffix = word[index+1:]
            
            if prefix.lower() in PREPOSITIONS and i > 0:
                 # Caso d'America (non è la prima parola, e la d' è una preposizione elisa)
                 formatted_word = prefix.lower() + suffix.capitalize()
            else:
                 # Caso L'Aquila / D'Annunzio (o è la prima parola, o non è una preposizione)
                 formatted_word = prefix.capitalize() + suffix.capitalize()

        else:
            # 3. Capitalizzazione standard (Vico, Roma, ecc.)
            formatted_word = word.capitalize()
            
        
        # Aggiungo la parola formattata
        final_formatted_words.append(formatted_word)
            
    return ' '.join(final_formatted_words)


# Incolliamo anche questa, può servire per la formattazione di nom_cog
def format_name_with_exceptions(name_str):
    """
    Formato una stringa di nome/cognome. Se non usi '%' la logica è Title Case.
    """
    if not name_str:
        return ""

    words = name_str.split()
    formatted_parts = []
    
    for word in words:
        if word.startswith('%'):
            cleaned_word = word[1:].lower()
            formatted_parts.append(cleaned_word)
        else:
            formatted_parts.append(word.lower().capitalize())
            
    return ' '.join(formatted_parts)

app = Flask(__name__, static_folder='static')

temp_pdf_batches = {}

CLEANUP_DELAY_SECONDS = 3600  # 1 ora

def cleanup_batch_data(batch_id):
    """Funzione per pulire i dati di un batch dopo un certo ritardo."""
    batch_info = temp_pdf_batches.pop(batch_id, None)
    if batch_info:
        try:
            shutil.rmtree(batch_info['temp_dir'])
            print(f"Pulita directory temporanea per batch {batch_id}: {batch_info['temp_dir']}")
        except Exception as e:
            print(f"Errore durante la pulizia della directory temporanea {batch_info['temp_dir']} per batch {batch_id}: {e}")

@app.route('/', methods=['GET'])
def homepage():
    return render_template('upload.html')

@app.route('/upload-data', methods=['POST'])
def upload_data():
    if 'data_file' not in request.files:
        return 'Nessun file caricato', 400

    file = request.files['data_file']
    if file.filename == '':
        return 'Nessun file selezionato', 400

    if file:
        file_content = file.stream.read().decode('utf-8')
        students_data = parse_diploma_data(file_content)

        if not students_data:
            return 'Impossibile leggere i dati dal file. Controlla il formato o che contenga dati validi.', 400

        data_creazione = datetime.now()
        nome_cartella = data_creazione.strftime('%Y-%m-%d')
        
        log_entries = []
        
        batch_id = str(uuid.uuid4())
        current_batch_temp_dir = tempfile.mkdtemp()
        
        generated_pdf_filenames = [] 

        # ... all'interno della funzione upload_data ...

        for i, student in enumerate(students_data):
             # Tutte le chiavi (incluse le nuove FIRMA4, LOGO1, etc.) sono messe in minuscolo.
            student_data_for_template = {
                key.lower(): value for key, value in student.items()
            }

            # Applicazione della formattazione avanzata al nome e cognome
            corsolau_raw = student_data_for_template.get('corsolau', '')
            student_data_for_template['corsolau'] = corsolau_raw.replace('|', '<br>')
            nom_cog_raw = student_data_for_template.get('nom_cog', '')
            student_data_for_template['nom_cog'] = format_name_with_exceptions(nom_cog_raw.replace('|', '<br>')) 

            # --- NUOVA LOGICA: Gestione e Formattazione di Luogo di Nascita e Stato (STATNAS) ---

            # Estrai e formatta i componenti separatamente
            comune_nascita = format_place_name(student_data_for_template.get('luogonas', '').strip())
            stato_nascita = format_place_name(student_data_for_template.get('statnas', '').strip())
            provincia_nascita = format_place_name(student_data_for_template.get('provnas', '').strip())

            # Ricostruisci la stringa finale
            luogo_nascita_completo = comune_nascita
            
            if provincia_nascita: 
                # Se c'è la provincia, la aggiungiamo tra parentesi (es. Cava de' Tirreni (Salerno))
                luogo_nascita_completo += f" ({provincia_nascita})"

            if stato_nascita and stato_nascita.upper() not in ['ITALIA', 'IT', 'I']: 
                # Se è uno Stato estero, lo aggiungiamo (e potresti voler decidere se rimuovere la provincia in questo caso)
                # Per semplicità, lo aggiungiamo come testo libero, ma potresti volerlo tra parentesi:
                luogo_nascita_completo += f" ({stato_nascita})" 

            # Sovrascriviamo la chiave 'luogonas' per l'uso nel template
            student_data_for_template['luogonas'] = luogo_nascita_completo
            student_data_for_template['statnas'] = stato_nascita # Aggiorna lo stato se dovesse servire separatamente
            student_data_for_template['provnas'] = provincia_nascita # Aggiorna la provincia
            
            # ----------------------------------------------------------------------

            modulo_value = student_data_for_template.get('modulo', '').strip()
            
            student_pdfs = []
            
            # --- INIZIO: Generazione del PDF del Diploma ---
            template_filename = ''
            
            # LOGICA DI SELEZIONE TEMPLATE AGGIORNATA
            if modulo_value == 'forml01v7':
                template_filename = 'diploma_forml01v7.html'
            elif modulo_value == 'forml01v7tuscia':
                template_filename = 'diploma_forml01v7tuscia.html'
            elif modulo_value == 'forml29v7':
                template_filename = 'diploma_forml29v7.html'
            elif modulo_value == 'forml28v7':
                template_filename = 'diploma_forml28v7.html'
            elif modulo_value == 'forml28v7A':
                template_filename = 'diploma_forml28v7A.html'
            elif modulo_value == 'memoriastudi':
                template_filename = 'diploma_memoriastudi.html'
            elif modulo_value == 'memorialaureamag':
                template_filename = 'diploma_memorialaureamag.html'
            elif modulo_value == 'memorialaureatri':
                template_filename = 'diploma_memorialaureatri.html'
            else:
                log_entries.append(f"ATTENZIONE: Modulo '{modulo_value}' non riconosciuto per {student_data_for_template.get('nom_cog', 'uno studente')}. Saltando la generazione del PDF.")
                continue 

            student_data_for_template['lode'] = student_data_for_template.get('lode', '').upper().strip()
            student_data_for_template['testo_footer_fisso'] = "Imposta di bollo assolta in modo virtuale. Autorizzazione Intendenza di Finanza di Roma n.9120/88"
            
            # Mappatura automatica delle chiavi FIRMA e LOGO in file .png
            # Questo gestisce le vecchie chiavi (firmar, firmap, firmad) e le nuove (firma4, firma5, firma6)
            for key in ['firmar', 'firmap', 'firmad', 'firma4', 'firma5', 'firma6']:
                val = student_data_for_template.get(key)
                if val and not val.endswith('.png'): # Evita di aggiungere .png se già presente (prevenzione)
                    student_data_for_template[key] = f"{val}.png"
                    
            # Mappatura automatica per i nuovi LOGHI
            for key in ['logo1', 'logo2', 'logo3']:
                val = student_data_for_template.get(key)
                if val and not val.endswith('.png'):
                    student_data_for_template[key] = f"{val}.png"
            
            # Esegui la generazione del template
            rendered_html = render_template(
                template_filename,
                **student_data_for_template
            )

            try:
                html_doc = HTML(string=rendered_html, base_url=request.url_root) ####nominazione file pdf
                pdf_bytes = html_doc.write_pdf()
                student_name_for_filename = student_data_for_template.get('nom_cog', f'studente_{i+1}').replace(' ', '_')
                if '<br>' in student_name_for_filename:
                    student_name_for_filename = student_data_for_template.get('nom_cog', f'studente_{i+1}').replace('<br>', '_')
                pdf_filename = f'diploma_{student_name_for_filename}_{modulo_value}.pdf'
                pdf_path = os.path.join(current_batch_temp_dir, pdf_filename)
                with open(pdf_path, 'wb') as f:
                    f.write(pdf_bytes)
                
                student_pdfs.append(pdf_filename)
                log_entries.append(f"Diploma generato per: {student_data_for_template.get('nom_cog', 'N/A')}")

            except Exception as e:
                print(f"Errore nella generazione del PDF per {student_data_for_template.get('nom_cog', 'uno studente')}: {e}")
                log_entries.append(f"ERRORE: Impossibile generare il PDF per {student_data_for_template.get('nom_cog', 'uno studente')}. Errore: {e}")
            
            # --- FINE: Generazione del PDF del Diploma ---
            
            # --- INIZIO: Generazione del PDF della Camicia ---
            camicia_data_for_template = {
                'corso_laurea': student_data_for_template.get('corsolau', ''),
                'nome_studente': student_data_for_template.get('nom_cog', ''),
                'luogo_nascita': student_data_for_template.get('luogonas', '').split('(')[0].strip() if student_data_for_template.get('luogonas') and '(' in student_data_for_template.get('luogonas') else student_data_for_template.get('luogonas', '').strip(),
                'provincia_nascita': student_data_for_template.get('provnas', '').split('(')[1].replace(')', '').strip() if '(' in student_data_for_template.get('provnas', '') else '',
                'data_nascita': student_data_for_template.get('datanas', ''),
                'data_stampa': student_data_for_template.get('datastamp', ''),
                'numero_protocollo': student_data_for_template.get('protocol', ''),
                'data_rilascio': student_data_for_template.get('datastamp', ''),
                'numero_diploma': student_data_for_template.get('npergamena', ''),
                'genere_nato_nata': student_data_for_template.get('sesso', 'nato a').strip(), 
                'classe_laurea_dinamica': student_data_for_template.get('indicorso', ''),
                'firmad': student_data_for_template.get('firmad', ''),
                'firmar': student_data_for_template.get('firmar', ''),
                'firmap': student_data_for_template.get('firmap', '')  
            }

            rendered_camicia_html = render_template(
                'camicia_template.html',
                **camicia_data_for_template
            )
            
            try:
                html_camicia_doc = HTML(string=rendered_camicia_html, base_url=request.url_root)
                pdf_camicia_bytes = html_camicia_doc.write_pdf()

                camicia_pdf_filename = f'camicia_{student_name_for_filename}.pdf'
                camicia_pdf_path = os.path.join(current_batch_temp_dir, camicia_pdf_filename)
                with open(camicia_pdf_path, 'wb') as f:
                    f.write(pdf_camicia_bytes)
                
                student_pdfs.append(camicia_pdf_filename)
                log_entries.append(f"Camicia generata per: {student_data_for_template.get('nom_cog', 'N/A')}")

            except Exception as e:
                print(f"Errore nella generazione della Camicia per {student_data_for_template.get('nom_cog', 'uno studente')}: {e}")
                log_entries.append(f"ERRORE: Impossibile generare la Camicia per {student_data_for_template.get('nom_cog', 'uno studente')}. Errore: {e}")
            # --- FINE: Generazione del PDF della Camicia ---

            # Aggiungi i nomi dei file generati per questo studente alla lista principale
            generated_pdf_filenames.extend(student_pdfs)

        # Recuperiamo metadati dal primo studente per l'archiviazione
        primo_studente = students_data[0] if students_data else {}
        
        # 1. Generazione del PDF combinato (solo diplomi)
        diploma_filenames = [f for f in generated_pdf_filenames if f.startswith('diploma_')]
        if diploma_filenames:
            merger = PdfWriter()
            combined_pdf_filename = f'tutti_i_diplomi_{nome_cartella}.pdf'
            combined_pdf_path = os.path.join(current_batch_temp_dir, combined_pdf_filename)
            for pdf_filename in diploma_filenames:
                file_path_to_merge = os.path.join(current_batch_temp_dir, pdf_filename)
                try:
                    merger.append(file_path_to_merge)
                except Exception as e:
                    print(f"Errore merge: {e}")
            merger.write(combined_pdf_path)
            merger.close()
            generated_pdf_filenames.append(combined_pdf_filename)

        # 2. Creazione del contenuto del LOG (ora lo definiamo PRIMA di usarlo)
        log_content = '\n'.join(log_entries)
        log_file_path = os.path.join(current_batch_temp_dir, 'log_creazione_diplomi.txt')
        with open(log_file_path, 'w', encoding='utf-8') as f:
            f.write(log_content)

        protocollo_raw = primo_studente.get('PROTOCOL', '').strip()
        protocollo_clean = protocollo_raw.split('/')[0] if '/' in protocollo_raw else protocollo_raw



        # 3. Preparazione METADATI per l'archiviazione (Prendiamo i dati dal primo studente del CSV)
        facolta = primo_studente.get('CORSOLAU', 'N-A').replace(' ', '')
        anno_laurea = primo_studente.get('DATALAUR', datetime.now().strftime('%Y')).replace(' ', '')
        tipologia = "Pergamena"
        # Esempio: se il modulo contiene 'tri', allora è Triennale
        modulo_primo = primo_studente.get('CLASSE', '')
        if 'LM-' in modulo_primo: tipologia = "LM"
        else: tipologia = "LT"

        # 4. Salvataggio nel dizionario globale (con log_content finalmente definito)
        temp_pdf_batches[batch_id] = {
            'temp_dir': current_batch_temp_dir,
            'filenames': generated_pdf_filenames,
            'log_content': log_content,
            'log_file_path': log_file_path,
            'original_folder_name': nome_cartella,
            'metadata': {
                'protocollo': protocollo_clean,
                'tipologia': tipologia,
                'facolta': facolta,
                'anno_laurea': anno_laurea,
                'nomi_persone': [s.get('NOM_COG', 'N/A') for s in students_data],
                'totale': len(students_data)
            }
        }
        
        # 5. Avvio timer pulizia
        timer = threading.Timer(CLEANUP_DELAY_SECONDS, cleanup_batch_data, args=[batch_id])
        timer.start()

        return redirect(url_for('preview_pdfs', batch_id=batch_id))
    
    return 'Errore sconosciuto', 500


#### Rotte di Servizio dei File
# --- NUOVA ROTTA ARCHIVIA ---

@app.route('/archive/<batch_id>', methods=['POST'])
def archive_batch(batch_id):
    batch_info = temp_pdf_batches.get(batch_id)
    if not batch_info:
        return "Batch non trovato", 404
    # Controllo duplicati
    if batch_info.get('archived', False):
        return "Questo batch è già stato archiviato.", 200

    meta = batch_info['metadata']
    now = datetime.now()
    timestamp_str = now.strftime('%d%m%Y_%H%M')
    
    # 1. Creazione Nome Cartella/Zip
    # Formato: LaureaTriennale_40_Ingegneria_2010_22122025_1530
    folder_name = f"{meta['tipologia']}_{meta['totale']}_{meta['facolta']}_{meta['anno_laurea']}_{timestamp_str}"
    zip_filename = f"{folder_name}.zip"
    temp_zip_path = os.path.join(batch_info['temp_dir'], zip_filename)

    # 2. Creazione del Log specifico per l'archivio
    archive_log_path = os.path.join(batch_info['temp_dir'], "log_archivio.txt")
    with open(archive_log_path, 'w', encoding='utf-8') as f:
        f.write(f"PRODUZIONE PERGAMENE - {now.strftime('%d/%m/%Y %H:%M')}\n")
        f.write(f"Tipologia: {meta['tipologia']}\n")
        f.write(f"Totale pergamene: {meta['totale']}\n")
        f.write("-" * 30 + "\n")
        for nome in meta['nomi_persone']:
            f.write(f"- {nome}\n")

    # 3. Creazione dello ZIP (contenente solo PDF e il nuovo log)
    with zipfile.ZipFile(temp_zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for filename in batch_info['filenames']:
            if filename.startswith('diploma_'): # Includiamo solo le pergamene come da tua richiesta
                file_path = os.path.join(batch_info['temp_dir'], filename)
                zf.write(file_path, arcname=os.path.join(folder_name, filename))
        zf.write(archive_log_path, arcname=os.path.join(folder_name, "log_nominativi.txt"))

    # 4. Copia nelle due cartelle locali
    try:
        shutil.copy2(temp_zip_path, os.path.join(PATH_ARCHIVIO_1, zip_filename))
        shutil.copy2(temp_zip_path, os.path.join(PATH_ARCHIVIO_2, zip_filename))
    except Exception as e:
        return f"Errore durante la copia nei server: {e}", 500

    # 5. Aggiornamento Registro Excel
    excel_path = os.path.join(PATH_EXCEL_REGISTRO, f"Pergamene_{now.year}.xlsx")
    new_row = [
                meta['protocollo'], 
                meta['tipologia'], 
                meta['totale'], 
                meta['facolta'], 
                meta['anno_laurea'], 
                now.strftime('%d/%m/%Y %H:%M')
            ]
    
    try:
        if not os.path.exists(excel_path):
            wb = Workbook()
            ws = wb.active
            ws.append(["Protocollo","Tipologia", "Totale PDF", "Facoltà", "Anno Laurea", "Data Stampa"])
        else:
            wb = load_workbook(excel_path)
            ws = wb.active
        
        ws.append(new_row)
        wb.save(excel_path)
        batch_info['archived'] = True
        return f"Archiviato con successo. Protocollo: {meta['protocollo']}", 200
    
    except Exception as e:
        return f"File Excel in uso o errore: {e}", 500

    #return f"Archiviazione completata con successo in: {folder_name}"

@app.route('/preview/<batch_id>')
def preview_pdfs(batch_id):
    batch_info = temp_pdf_batches.get(batch_id)
    if not batch_info:
        return "Anteprima non trovata o scaduta.", 404

    pdf_list_for_template = []
    # Prepara la lista di PDF per il template, includendo SOLO i diplomi
    for filename in batch_info['filenames']:
        # Aggiungi questa condizione per filtrare solo i PDF dei diplomi
        if filename.startswith('diploma_'): 
            pdf_list_for_template.append({
                'name': filename,
                'url': url_for('get_single_pdf', batch_id=batch_id, filename=filename)
            })

    return render_template('preview.html',
                            pdf_list=pdf_list_for_template,
                            download_url=url_for('download_zip_for_preview', batch_id=batch_id),
                            log_url=url_for('get_log_for_preview', batch_id=batch_id),
                            cleanup_delay_minutes=CLEANUP_DELAY_SECONDS / 60)

@app.route('/preview/pdf/<batch_id>/<filename>')
def get_single_pdf(batch_id, filename):
    batch_info = temp_pdf_batches.get(batch_id)
    if not batch_info:
        return "File non trovato.", 404
    
    if filename not in batch_info['filenames']:
        return "File non autorizzato o non trovato nel batch.", 403

    return send_file(os.path.join(batch_info['temp_dir'], filename), mimetype='application/pdf')

@app.route('/preview/log/<batch_id>')
def get_log_for_preview(batch_id):
    batch_info = temp_pdf_batches.get(batch_id)
    if not batch_info:
        return "Log non trovato o scaduto.", 404
    
    return send_file(batch_info['log_file_path'], 
                    mimetype='text/plain', 
                    as_attachment=True, 
                    download_name='log_creazione_diplomi.txt')


@app.route('/download_zip/<batch_id>')
def download_zip_for_preview(batch_id):
    batch_info = temp_pdf_batches.get(batch_id)
    if not batch_info:
        return "Download non trovato o scaduto.", 404

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'a', zipfile.ZIP_DEFLATED, False) as zf:
        for filename in batch_info['filenames']:
            file_path = os.path.join(batch_info['temp_dir'], filename)
            
            if filename.startswith('diploma_'):
                subfolder = 'pergamene'
            elif filename.startswith('camicia_'):
                subfolder = 'camicie'
            elif filename.startswith('tutti_i_diplomi_'):
                subfolder = 'combinato'
            else:
                subfolder = 'altri' 

            arcname = os.path.join(batch_info['original_folder_name'], subfolder, filename)
            zf.write(file_path, arcname=arcname)
        
        log_filename_in_zip = os.path.join(batch_info['original_folder_name'], 'log_creazione_documenti.txt')
        zf.write(batch_info['log_file_path'], arcname=log_filename_in_zip)
    
    zip_buffer.seek(0)
    
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'documenti_{batch_info["original_folder_name"]}.zip'
    )

def parse_diploma_data(file_content):
    lines = file_content.splitlines()
    if len(lines) < 4:
        return []

    header_line = lines[3]
    data_lines = lines[4:]

    reader = csv.reader(io.StringIO(header_line + '\n' + '\n'.join(data_lines)), delimiter='^')

    try:
        headers = next(reader)
    except StopIteration:
        return []

    students_data = []
    for row in reader:
        if row and len(row) == len(headers):
            student_dict = {}
            for i, header in enumerate(headers):
                student_dict[header.strip()] = row[i].strip()
            students_data.append(student_dict)
        else:
            pass

    return students_data

if __name__ == '__main__':
    app.run(debug=True)